from datetime import datetime
import io
import math
import re
from urllib import request
import azure.functions as func
import json
import logging
import os
import pdfplumber
from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient, ContentSettings
from openai import AzureOpenAI
from azure.storage.blob import BlobClient
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta ,date
import requests
import snowflake.connector
import sys
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.hazmat.primitives.asymmetric import dsa
from cryptography.hazmat.primitives import serialization


# ---------------- Load .env ----------------
load_dotenv()
# ---------------- Logging ----------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)


AZURE_OPENAI_ENDPOINT = os.getenv("ENDPOINT_URL")
AZURE_OPENAI_KEY = os.getenv("AZURE_OPENAI_API_KEY")
DEPLOYMENT_NAME = os.getenv("DEPLOYMENT_NAME", "gpt-4.1")

SEARCH_ENDPOINT = os.getenv("SEARCH_ENDPOINT")
SEARCH_KEY = os.getenv("SEARCH_KEY")
SEARCH_INDEX = os.getenv("SEARCH_INDEX_NAME")

client = AzureOpenAI(
    azure_endpoint=AZURE_OPENAI_ENDPOINT,
    api_key=AZURE_OPENAI_KEY,
    api_version="2025-01-01-preview",
)


# ---------------- Function App ----------------
app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)



@app.route(
    route="projects/{projectName}/upload",
    methods=["POST"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def upload_project_file(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Blob upload triggered")

    try:
        project_name = req.route_params.get("projectName")
        file = req.files.get("file")

        if not project_name or not file:
            return func.HttpResponse(
                json.dumps({"error": "projectName and file are required"}),
                status_code=400,
                mimetype="application/json"
            )

        # Sanitize project name
        project_name = project_name.replace("..", "").replace("/", "_")

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")

        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)

        blob_path = f"{project_name}/{file.filename}"

        blob_client = container_client.get_blob_client(blob_path)

        blob_client.upload_blob(
            file.stream.read(),
            overwrite=True,
            content_settings=ContentSettings(
                content_type=file.content_type
            )
        )

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "project": project_name,
                "file": file.filename,
                "path": blob_path
            }),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Blob upload failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

@app.route(
    route="projects",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def list_projects(req: func.HttpRequest) -> func.HttpResponse:
    try:
        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")

        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)

        projects = set()
        for blob in container_client.list_blobs():
            projects.add(blob.name.split("/")[0])

        return func.HttpResponse(
            json.dumps(sorted(projects)),
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Failed to list projects")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(
    route="projects/{projectName}/files",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def list_project_files(req: func.HttpRequest) -> func.HttpResponse:
    try:
        project_name = req.route_params.get("projectName")

        if not project_name:
            return func.HttpResponse(
                json.dumps({"error": "projectName required"}),
                status_code=400,
                mimetype="application/json"
            )

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")

        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)

        files = []
        prefix = f"{project_name}/"

        for blob in container_client.list_blobs(name_starts_with=prefix):
            files.append(blob.name.replace(prefix, ""))

        return func.HttpResponse(
            json.dumps(files),
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Failed to list project files")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )
    

def append_to_final_report(project: str, entry_type: str, payload: dict):
    excel_path = f"daily-reports/{project}/final-report.xlsx"
    blob_client = get_blob_client(excel_path)

    output = io.BytesIO()

    try:
        existing = io.BytesIO()
        blob_client.download_blob().readinto(existing)
        existing.seek(0)
        wb = load_workbook(existing)
        ws = wb.active
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "logs"
        ws.append(["date", "type", "data"])

    ws.append([
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        entry_type,
        json.dumps(payload)
    ])

    wb.save(output)
    output.seek(0)

    blob_client.upload_blob(
        output,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )


@app.route(
    route="projects/daily-reports/{projectName}/upload",
    methods=["POST"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def upload_daily_report(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Daily report upload triggered")

    try:
        project_name = req.route_params.get("projectName")
        file = req.files.get("file")

        if not project_name or not file:
            return func.HttpResponse(
                json.dumps({"error": "projectName and file are required"}),
                status_code=400,
                mimetype="application/json"
            )

        project_name = project_name.replace("..", "").replace("/", "_")

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")

        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)

        blob_path = f"daily-reports/{project_name}/{file.filename}"
        blob_client = container_client.get_blob_client(blob_path)

        # 1️⃣ Upload PDF
        blob_client.upload_blob(
            file.stream.read(),
            overwrite=True,
            content_settings=ContentSettings(
                content_type=file.content_type
            )
        )

        # 2️⃣ Extract PDF text
        extracted_text = read_pdf_from_blob_path(blob_path)

        # 3️⃣ Prepare payload
        payload = {
            "project": project_name,
            "fileName": file.filename,
            "uploadedAt": datetime.utcnow().isoformat(),
            "type": "daily_report",
            "content": extracted_text
        }

        # 4️⃣ Append to final-report.xlsx
        append_to_final_report(
            project=project_name,
            entry_type="daily_report",
            payload=payload
        )

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "project": project_name,
                "file": file.filename,
                "path": blob_path
            }),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Daily report upload failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

def read_pdf_from_blob(project:str,blob_name: str) -> str:
    blob_service = BlobServiceClient.from_connection_string(os.getenv("AZURE_STORAGE_CONNECTION_STRING"))
    full_blob_path = f"{project}/{blob_name}"
    blob_client = blob_service.get_blob_client(container=os.getenv("BLOB_CONTAINER_NAME"), blob=full_blob_path)
    logging.info(blob_name)
    
    stream = io.BytesIO()
    blob_client.download_blob().readinto(stream)
    
    text = ""
    with pdfplumber.open(stream) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

@app.route(route="contracts/compare", methods=["POST"])
def compare_reports(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Daily-reports comparison triggered")

    try:
        body = req.get_json()

        project = body.get("projectName")
        files = body.get("files", [])

        if not project or len(files) < 2:
            return func.HttpResponse(
                json.dumps({"error": "At least two files are required for comparison."}),
                status_code=400,
                mimetype="application/json"
            )

        logging.info(f"Comparing files | Project={project} | Files={files}")

        file_texts = []
        for file in files:
            file_text = read_pdf_from_blob(project, file)
            file_texts.append(file_text)

        # --- Prompt Engineering ---
        prompt = f"""
You are a senior construction contract analyst with expertise in interior construction projects.

Project: {project}

Compare the following contracts in a concise, executive-friendly format. For each point, include a citation referring to the section or line in the provided PDF from which the summary is taken.

Instructions:
1. Start with a **Final Recommendation**: which contract is preferable and why.
2. Immediately follow with **Reasons** (bullet points, concise).
3. Present **Key Differences** in a **table format** using the actual file names as headers:
   | Aspect       | {" | ".join(files)} |
   |-------------|-----------|
4. Cover the following aspects in the table: Scope, Commercials, Timelines, Risks.
5. For each bullet point or table entry, reference the PDF section, e.g., "Scope of Work – Section 1" or "Weekly Execution Timeline – Section 3".
6. Avoid long paragraphs—use bullet points and tables for clarity.
7. Always reference the actual file names in all sections.

Contracts:
"""
        for i, file_text in enumerate(file_texts):
            prompt += f"Contract {i + 1}: {files[i]}\nContent:\n{{file_text}}\n\n"

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You compare construction contracts."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.2
        )

        comparison_text = completion.choices[0].message.content

        logging.info("Comparison generated successfully")

        return func.HttpResponse(
            json.dumps({
                "comparison": comparison_text
            }),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Comparison failed")

        return func.HttpResponse(
            json.dumps({
                "error": str(e)
            }),
            status_code=500,
            mimetype="application/json"
        )
    
def read_pdf_from_blob_path(blob_path: str) -> str:
    blob_service = BlobServiceClient.from_connection_string(
        os.getenv("AZURE_STORAGE_CONNECTION_STRING")
    )

    container_name = os.getenv("BLOB_CONTAINER_NAME")
    blob_client = blob_service.get_blob_client(
        container=container_name,
        blob=blob_path
    )

    stream = io.BytesIO()
    blob_client.download_blob().readinto(stream)
    stream.seek(0)

    text = ""
    with pdfplumber.open(stream) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

    return text


def calculate_reporting_week(start_date: str, report_date: str) -> int:
    start = datetime.strptime(start_date, "%Y-%m-%d")
    report = datetime.strptime(report_date, "%Y-%m-%d")

    delta_days = (report - start).days
    if delta_days < 0:
        return 1

    return math.floor(delta_days / 7) + 1

def extract_report_date(report_text: str) -> str | None:
    """
    Extracts date in formats like:
    14-Jan-2026, 14 Jan 2026, 2026-01-14
    """
    patterns = [
        r"\b\d{1,2}-[A-Za-z]{3}-\d{4}\b",
        r"\b\d{1,2}\s[A-Za-z]{3}\s\d{4}\b",
        r"\b\d{4}-\d{2}-\d{2}\b",
    ]

    for pattern in patterns:
        match = re.search(pattern, report_text)
        if match:
            raw = match.group(0)
            for fmt in ("%d-%b-%Y", "%d %b %Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
                except:
                    continue

    return None




@app.route(route="daily-reports/anomaly-detect", methods=["POST"])
def detect_daily_report_anomalies(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Daily report anomaly detection triggered")

    try:
        body = req.get_json()

        project = body.get("projectName")
        files = body.get("files", [])
        start_date = body["anomalyStartDate"]

        daily_report_file = files[0]
        final_sow_file = files[1]

        # ---------- READ DOCUMENTS ----------
        daily_report_text_full = read_pdf_from_blob_path(
            f"daily-reports/{project}/{daily_report_file}"
        )

        final_sow_text = read_pdf_from_blob(project, final_sow_file)

        # ---------- EXTRACT REPORT DATE ----------
        daily_report_date = extract_report_date(daily_report_text_full)

        if not daily_report_date:
            raise ValueError("Could not extract Daily Report date from document")

        # ---------- CALCULATE REPORTING WEEK ----------
        reporting_week = calculate_reporting_week(
            start_date=start_date,
            report_date=daily_report_date
        )

        daily_report_text = daily_report_text_full
        # ---------- 🔥 DYNAMIC PROMPT ----------
        prompt = f"""
You are a senior construction controls and contract compliance analyst.

Project: {project}

CONTEXT (AUTHORITATIVE):
- Project start date: {start_date}
- Daily report date: {daily_report_date}
- Calculated reporting week: Week {reporting_week}

STRICT EVALUATION RULES:
1. Only assess activities that were contractually scheduled
   to START or PROGRESS in Week {reporting_week} or earlier.
2. Any activity scheduled AFTER Week {reporting_week}:
   - MUST be Impact: Low
   - Use wording: "Not yet due as per contract timeline"
3. Medium or High impact is allowed ONLY if:
   - The activity was due by Week {reporting_week}
   - AND the Daily Report shows a clear deviation
4. Payment-related deviations:
   - Medium impact ONLY if the linked physical milestone
     was due by Week {reporting_week}
   - Otherwise Impact: Low

OUTPUT FORMAT (MANDATORY — UI DEPENDS ON THIS):
• Category | Expected | Observed | Impact: High / Medium / Low

STYLE RULES:
- One bullet per deviation
- Short, factual statements only
- No explanations
- No headings
- No extra text before or after bullets

FINAL SOW (Baseline):
{final_sow_text}

DAILY REPORT (Week {reporting_week}):
{daily_report_text}

If NO valid deviations exist, respond with EXACTLY:
No deviations detected for the current reporting period.
"""

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You detect construction compliance anomalies."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        anomaly_report = completion.choices[0].message.content

        return func.HttpResponse(
            json.dumps({
                "anomalies": anomaly_report,
                "reportingWeek": reporting_week,
                "reportDate": daily_report_date
            }),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Anomaly detection failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(route="document-chat", methods=["POST"])
def document_chat(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Document chat triggered")

    try:
        body = req.get_json()

        project = body.get("projectName")
        file_names = body.get("fileNames") or body.get("fileName")  # Support both single and multiple
        question = body.get("question")

        # Handle both single file (string) and multiple files (list)
        if isinstance(file_names, str):
            file_names = [file_names]

        if not file_names or not question or not project:
            return func.HttpResponse(
                json.dumps({"error": "projectName, fileNames (or fileName), and question are required."}),
                status_code=400,
                mimetype="application/json"
            )

        # Read all PDF contents from blob
        documents_content = []
        for file_name in file_names:
            file_text = read_pdf_from_blob(project, file_name)
            documents_content.append(f"DOCUMENT: {file_name}\nContent:\n{file_text}\n")

        all_documents = "\n\n---\n\n".join(documents_content)

        # ---------- 🔥 NEW PROMPT (CHAT BASED ON DOCUMENT(S)) ----------
        document_label = "document" if len(file_names) == 1 else "documents"
        file_list = ", ".join(file_names)
        
        prompt = f"""
You are a senior construction project assistant with expert knowledge in reading and analyzing construction project documents. 

Project: {project}

TASK:
Answer the user's question based ONLY on the content of the provided {document_label}.
Do NOT provide any information that is not in the {document_label}.
If the question requires analysis or recommendations (e.g., vendor selection), base your reasoning strictly on the document content.
When comparing multiple documents, clearly reference which document supports each point.

DOCUMENTS: {file_list}

{all_documents}

USER QUESTION:
{question}

RESPONSE REQUIREMENTS:
- Be precise, factual, and detailed.
- Summarize content when asked for a summary.
- If asked for recommendations (e.g., which vendor to choose), provide reasoning strictly based on the document's information.
- When answering from multiple documents, cite which document each piece of information comes from.
- Keep your answer concise but comprehensive.
- Provide your answer only in paragraph form (no tables).
- If the answer is not in the {document_label}, respond with:
  "The {document_label} do not contain information to answer this question."
"""

        # Call the OpenAI chat model
        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You answer user questions based on provided documents."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        answer = completion.choices[0].message.content
        logging.info(f"Document chat response: {answer}")

        return func.HttpResponse(
            json.dumps({"answer": answer}),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Document chat failed")

        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )
    

@app.route(
    route="projects/{projectName}/files",
    methods=["DELETE"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def delete_file(req: func.HttpRequest) -> func.HttpResponse:
    try:
        project_name = req.route_params.get("projectName")
        file_name = req.params.get("fileName")

        if not project_name or not file_name:
            return func.HttpResponse(
                json.dumps({"error": "projectName and fileName required"}),
                status_code=400,
                mimetype="application/json"
            )

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")

        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)

        blob_path = f"{project_name}/{file_name}"
        blob_client = container_client.get_blob_client(blob_path)

        blob_client.delete_blob()

        return func.HttpResponse(
            json.dumps({"message": f"{file_name} deleted successfully"}),
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Failed to delete file")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )
    

@app.route(
    route="projects/{projectName}/files/meta",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def list_project_files_with_metadata(req: func.HttpRequest) -> func.HttpResponse:
    """Return a list of files in the given project with metadata (last_modified/upload date, size, content type, metadata)."""
    try:
        project_name = req.route_params.get("projectName")
 
        if not project_name:
            return func.HttpResponse(
                json.dumps({"error": "projectName required"}),
                status_code=400,
                mimetype="application/json"
            )
 
        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("BLOB_CONTAINER_NAME")
 
        blob_service = BlobServiceClient.from_connection_string(conn_str)
        container_client = blob_service.get_container_client(container_name)
 
        files = []
        prefix = f"daily-reports/{project_name}/"
 
        for blob in container_client.list_blobs(name_starts_with=prefix):
            # blob.name is the full path in the container
            file_name = blob.name.replace(prefix, "")
            if not file_name:
                # skip directory-like blobs
                continue
 
            # last_modified is a timezone-aware datetime if present
            last_modified = None
            if getattr(blob, "last_modified", None):
                try:
                    last_modified = blob.last_modified.isoformat()
                except Exception:
                    last_modified = str(blob.last_modified)
 
            size = getattr(blob, "size", None)
 
            content_type = None
            if getattr(blob, "content_settings", None):
                content_type = getattr(blob.content_settings, "content_type", None)
 
            files.append({
                "name": file_name,
                "last_modified": last_modified,
                "content_type": content_type,
            })
 
        return func.HttpResponse(
            json.dumps(files),
            mimetype="application/json"
        )
 
    except Exception as e:
        logging.exception("Failed to list project files with metadata")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

def get_blob_client(blob_path: str):
    return BlobServiceClient.from_connection_string(
        os.getenv("AZURE_STORAGE_CONNECTION_STRING")
    ).get_blob_client(
        container=os.getenv("BLOB_CONTAINER_NAME"),
        blob=blob_path
    )


def get_latest_final_report_payload(project: str) -> dict:
    excel_path = f"daily-reports/{project}/final-report.xlsx"
    blob_client = get_blob_client(excel_path)

    stream = io.BytesIO()
    blob_client.download_blob().readinto(stream)
    stream.seek(0)

    wb = load_workbook(stream)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        raise ValueError("No rows found in final-report.xlsx")

    headers = [str(h).lower() for h in rows[0]]
    try:
        type_idx = headers.index("type")
        data_idx = headers.index("data")
        date_idx = headers.index("date")
    except ValueError:
        raise ValueError("final-report.xlsx missing required columns")

    # Find latest row where type == 'final' AND fileName contains 'Anomaly_Version'
    latest_row = None
    payload = None
    
    for row in reversed(rows[1:]):
        if str(row[type_idx]).lower() == "final":
            # Parse the payload to check fileName
            payload_raw = row[data_idx]
            try:
                if isinstance(payload_raw, str):
                    temp_payload = json.loads(payload_raw)
                elif isinstance(payload_raw, dict):
                    temp_payload = payload_raw
                else:
                    temp_payload = json.loads(json.dumps(payload_raw, default=str))
                
                file_name = temp_payload.get("fileName", "")
                
                # Only accept rows with Anomaly_Version in fileName
                if "Anomaly_Version" in file_name:
                    latest_row = row
                    payload = temp_payload
                    logging.info(f"Found final report with Anomaly_Version: {file_name}")
                    break
                else:
                    logging.info(f"Skipping final entry without Anomaly_Version: {file_name}")
            except Exception as e:
                logging.warning(f"Failed to parse payload for row, skipping: {str(e)}")
                continue

    if not latest_row or not payload:
        raise ValueError("No final entry with Anomaly_Version found in final-report.xlsx")

    file_name = payload.get("fileName", "")
    logging.info(f"Final Report File Name: {file_name}")
    logging.info(f"Final Report Date: {latest_row[date_idx]}")

    return {
        "payload": payload,
        "date": latest_row[date_idx]
    }


def analyze_po_against_final(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("PO analyze triggered")

    try:
        project = req.route_params.get("projectName")
        body = req.get_json()
        po = body.get("purchaseOrder")

        if not project or not po:
            return func.HttpResponse(
                json.dumps({"error": "projectName and purchaseOrder are required"}),
                status_code=400,
                mimetype="application/json"
            )

        latest_final = get_latest_final_report_payload(project)
        final_content = latest_final["payload"].get("content", "")

        prompt = f"""
You are a senior construction controls analyst.

TASK:
Compare the Purchase Order (PO) against the FINAL report content (SOW) and identify anomalies.

PO DATA (JSON):
{json.dumps(po)}

FINAL REPORT CONTENT:
{final_content}

OUTPUT FORMAT (STRICT JSON ARRAY):
[
  {{
    "id": "string",
    "type": "quantity_mismatch | price_mismatch | total_mismatch | scope_mismatch | timeline_mismatch | compliance_gap",
    "severity": "low | medium | high",
        "description": "ITEM NAME ONLY (no extra text)",
    "field": "string",
    "expectedValue": "string or number",
    "actualValue": "string or number",
    "step": "po_receipt"
  }}
]

RULES - CRITICAL:
- Return ONLY valid JSON (no markdown).
- If no anomalies, return EMPTY ARRAY: []
- Use step="po_receipt" so the frontend labels it as Step 1.
- Treat FINAL REPORT CONTENT as the SOW source of truth.
- ONLY flag actual mismatches where expectedValue ≠ actualValue.
- Do NOT flag items when expected == actual (they match perfectly).
- Only compare PO items to SOW items that match by description or equivalent name.
- For matched items, ONLY return an anomaly if quantity or price differs.
- If a PO item does NOT exist in SOW, return an anomaly (type: scope_mismatch).
- Do NOT flag SOW-only items that are not present in the PO.
- Do NOT compare or mention client/vendor names, dates, totals, project names, or any non-line-item fields.
- Be VERY STRICT: Only return an anomaly if there is a CLEAR NUMERIC MISMATCH.
- Do NOT return anomalies for items where values match exactly.

SEVERITY RULES:
- type: price_mismatch → severity: "high" (always)
- type: total_mismatch → severity: "high" (always)
- type: quantity_mismatch where PO_quantity > SOW_quantity → severity: "high" (ordering more than needed)
- type: quantity_mismatch where PO_quantity < SOW_quantity → severity: "medium" (ordering less than needed)
- type: scope_mismatch (PO item not in SOW) → severity: "high"
- if the quantity in PO is less than quantity in SOW -> severity: "medium"
"""

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You analyze PO vs final report and return anomalies as JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        anomalies_raw = completion.choices[0].message.content.strip()
        anomalies = json.loads(anomalies_raw)

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "anomalies": anomalies
            }),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("PO analyze failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(route="projects/{projectName}/receipt-analyze", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def analyze_receipt_against_po(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Receipt analyze triggered")

    try:
        project = req.route_params.get("projectName")
        body = req.get_json()
        po = body.get("purchaseOrder")
        receipt = body.get("receipt")

        if not project or not po or not receipt:
            return func.HttpResponse(
                json.dumps({"error": "projectName, purchaseOrder, and receipt are required"}),
                status_code=400,
                mimetype="application/json"
            )

        prompt = f"""
You are a senior construction controls analyst.

TASK:
Compare the Receipt against the Purchase Order (PO) and identify anomalies.
The PO is the source of truth.

PO DATA (JSON):
{json.dumps(po)}

RECEIPT DATA (JSON):
{json.dumps(receipt)}

OUTPUT FORMAT (STRICT JSON ARRAY):
[
  {{
    "id": "string",
    "type": "quantity_mismatch | price_mismatch | total_mismatch | missing_item | extra_item",
    "severity": "low | medium | high",
        "description": "ITEM NAME ONLY (no extra text)",
    "field": "string",
    "expectedValue": "string or number",
    "actualValue": "string or number",
    "step": "po_receipt"
  }}
]

RULES - CRITICAL:
- Return ONLY valid JSON (no markdown).
- If no anomalies, return EMPTY ARRAY: []
- Use step="po_receipt" so the frontend labels it as Step 1.
- Treat PO as the source of truth.
- ONLY flag actual mismatches where expectedValue ≠ actualValue.
- Do NOT flag items when expected == actual (they match perfectly).
- Match items by description or equivalent name (case-insensitive).
- If a PO item does NOT exist in the receipt, return a missing_item anomaly.
- If a receipt item does NOT exist in the PO, return an extra_item anomaly.
- For matched items, ONLY return an anomaly if quantity or unit price differs.
- Do NOT compare or mention vendor/client names, dates, totals, project names, or any non-line-item fields.
- Be VERY STRICT: Only return an anomaly if there is a CLEAR NUMERIC MISMATCH.
- Do NOT return anomalies for items where values match exactly.
- Set description to ONLY the item name (e.g., Route Survey & Geotechnical Boreholes). Do not include any other text.

SEVERITY RULES:
- type: price_mismatch → severity: "high" (always)
- type: total_mismatch → severity: "high" (always)
- type: quantity_mismatch where Receipt_quantity > PO_quantity → severity: "high" (received more than ordered)
- type: quantity_mismatch where Receipt_quantity < PO_quantity → severity: "medium" (received less than ordered)
- type: missing_item → severity: "high"
- type: extra_item → severity: "high"
"""

        logging.info("Receipt analyze input PO JSON:\n%s", json.dumps(po, indent=2, ensure_ascii=False))
        logging.info("Receipt analyze input Receipt JSON:\n%s", json.dumps(receipt, indent=2, ensure_ascii=False))

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You analyze Receipt vs PO and return anomalies as JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        anomalies_raw = completion.choices[0].message.content.strip()
        logging.info("Receipt analyze model response (raw): %s", anomalies_raw)
        anomalies = json.loads(anomalies_raw)

        # Get receipt ID from receipt data
        receipt_id = receipt.get("id") or receipt.get("receiptId") or receipt.get("receiptNumber")
        order_id = receipt.get("poNumber") or po.get("OrderID") or po.get("orderId")

        # Save anomalies to Snowflake if we have receipt ID and order ID
        if receipt_id and order_id:
            try:
                ctx = connection_snowflake("CONSTRUCTIONDEMO", "partner", "SCT_SARORA_DB")
                cursor = ctx.cursor()

                # Prepare anomalies JSON with timestamps
                anomalies_with_timestamps = []
                for anomaly in anomalies:
                    anomaly_entry = {
                        "id": anomaly.get("id"),
                        "type": anomaly.get("type"),
                        "severity": anomaly.get("severity"),
                        "description": anomaly.get("description"),
                        "field": anomaly.get("field"),
                        "expectedValue": anomaly.get("expectedValue"),
                        "actualValue": anomaly.get("actualValue"),
                        "step": anomaly.get("step"),
                        "created_at": datetime.utcnow().isoformat(),
                        "updated_at": datetime.utcnow().isoformat()
                    }
                    anomalies_with_timestamps.append(anomaly_entry)

                if len(anomalies_with_timestamps) == 0:
                    update_sql = (
                        "UPDATE RECEIPTS "
                        "SET ANOMALIES = NULL, STATUS = %s, UPDATED_AT = CURRENT_TIMESTAMP() "
                        "WHERE RECEIPT_ID = %s"
                    )
                    cursor.execute(update_sql, ["submitted", receipt_id])
                else:
                    anomalies_json = json.dumps(anomalies_with_timestamps)
                    update_sql = (
                        "UPDATE RECEIPTS "
                        "SET ANOMALIES = TRY_PARSE_JSON(%s), UPDATED_AT = CURRENT_TIMESTAMP() "
                        "WHERE RECEIPT_ID = %s"
                    )
                    cursor.execute(update_sql, [anomalies_json, receipt_id])
                ctx.commit()

                logging.info(f"Saved {len(anomalies_with_timestamps)} anomalies to RECEIPTS for receipt {receipt_id}")

                disconnect_snowflake(ctx)
            except Exception as e:
                logging.warning(f"Failed to save anomalies to Snowflake: {str(e)}")

        response_data = {"success": True}
        if len(anomalies) > 0:
            response_data["anomalies"] = anomalies

        return func.HttpResponse(
            json.dumps(response_data),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Receipt analyze failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(route="projects/{projectName}/invoice-analyze-and-sync", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def analyze_invoice_against_receipt_and_sync(req: func.HttpRequest) -> func.HttpResponse:
    """
    Analyze Invoice against Receipt and sync to Salesforce.
    
    Request body:
    {
        "receipt": { receipt object },
        "invoice": { invoice object },
        "orderId": "order-123"
    }
    """
    logging.info("Invoice analyze and sync to Salesforce triggered")

    try:
        project = req.route_params.get("projectName")
        body = req.get_json()
        receipt = body.get("receipt")
        invoice = body.get("invoice")
        order_id = body.get("orderId")

        if not project or not receipt or not invoice or not order_id:
            return func.HttpResponse(
                json.dumps({"error": "projectName, receipt, invoice, and orderId are required"}),
                status_code=400,
                mimetype="application/json"
            )

        logging.info(f"Project: {project}, OrderID: {order_id}")
        logging.info("Receipt analyze input Receipt JSON:\n%s", json.dumps(receipt, indent=2, ensure_ascii=False))
        logging.info("Invoice analyze input Invoice JSON:\n%s", json.dumps(invoice, indent=2, ensure_ascii=False))

        prompt = f"""
You are a senior construction controls analyst.

TASK:
Compare the Invoice against the Receipt and identify anomalies.
The Receipt is the source of truth.

RECEIPT DATA (JSON):
{json.dumps(receipt)}

INVOICE DATA (JSON):
{json.dumps(invoice)}

OUTPUT FORMAT (STRICT JSON ARRAY):
[
  {{
    "id": "string",
    "type": "quantity_mismatch | price_mismatch | total_mismatch | missing_item | extra_item",
    "severity": "low | medium | high",
    "description": "ITEM NAME ONLY (no extra text)",
    "field": "string",
    "expectedValue": "string or number",
    "actualValue": "string or number",
    "step": "invoice_receipt"
  }}
]

RULES - CRITICAL:
- Return ONLY valid JSON (no markdown).
- If no anomalies, return EMPTY ARRAY: []
- Use step="invoice_receipt" for the frontend.
- Treat RECEIPT as the source of truth.
- ONLY flag actual mismatches where expectedValue ≠ actualValue.
- Do NOT flag items when expected == actual (they match perfectly).
- Match items by description or equivalent name (case-insensitive).
- If a receipt item does NOT exist in the invoice, return a missing_item anomaly.
- If an invoice item does NOT exist in the receipt, return an extra_item anomaly.
- For matched items, ONLY return an anomaly if quantity or unit price differs.
- Do NOT compare or mention vendor/client names, dates, totals, project names, or any non-line-item fields.
- Be VERY STRICT: Only return an anomaly if there is a CLEAR NUMERIC MISMATCH.
- Do NOT return anomalies for items where values match exactly.
- Set description to ONLY the item name (e.g., Mainline Ball Valves). Do not include any other text.

SEVERITY RULES:
- type: price_mismatch → severity: "high" (always)
- type: total_mismatch → severity: "high" (always)
- type: quantity_mismatch where Invoice_quantity > Receipt_quantity → severity: "high" (invoicing more than received)
- type: quantity_mismatch where Invoice_quantity < Receipt_quantity → severity: "medium" (invoicing less than received)
- type: missing_item → severity: "high"
- type: extra_item → severity: "high"
"""

        logging.info("Invoice analyze prompt:\n%s", prompt)

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You analyze Invoice vs Receipt and return anomalies as JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        anomalies_raw = completion.choices[0].message.content.strip()
        logging.info("Invoice analyze model response (raw): %s", anomalies_raw)
        anomalies = json.loads(anomalies_raw)

        logging.info(f"Parsed Invoice Anomalies: {json.dumps(anomalies, indent=2)}")

        # Get Salesforce token
        sf_auth = get_salesforce_access_token()
        access_token = sf_auth["access_token"]
        instance_url = sf_auth["instance_url"]

        # Convert anomalies to Salesforce format
        salesforce_anomalies = []
        
        for anomaly in anomalies:
            severity = (anomaly.get("severity") or "").lower()
            severity_label = {
                "high": "High Severity",
                "medium": "Medium Severity",
                "low": "Low Severity",
            }.get(severity, "Medium Severity")

            # Include field type in description for frontend parsing
            description = anomaly.get("description") or anomaly.get("type") or "Issue"
            field_type = anomaly.get("field", "amount")
            detail_with_type = f"{description} / {field_type}"

            salesforce_anomalies.append({
                "Detail": detail_with_type,
                "Severity": severity_label,
                "Expected": anomaly.get("expectedValue"),
                "Actual": anomaly.get("actualValue"),
                "Status": "Open"
            })

        # Determine invoice status
        invoice_status = "To be Paid" if len(anomalies) == 0 else "To be Verified"

        payload = {
            "OrderID": order_id,
            "InvoiceNumber": invoice.get("invoiceNumber") or invoice.get("InvoiceNumber") or "",
            "InvoiceStatus": invoice_status,
            "Anomalies": salesforce_anomalies
        }

        # Wrap payload in array for Salesforce API
        payload_array = [payload]
        
        logging.info(f"Invoice Sync Payload: {json.dumps(payload_array, indent=2)}")

        # Call Salesforce API to sync invoice
        salesforce_url = instance_url.rstrip("/") + "/services/apexrest/Invoice"
        logging.info(f"Calling Salesforce Invoice API: {salesforce_url}")

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        resp = requests.patch(salesforce_url, json=payload_array, headers=headers)
        
        # Log response for debugging
        logging.info(f"Salesforce Invoice API Response Status: {resp.status_code}")
        logging.info(f"Salesforce Invoice API Response: {resp.text}")
        
        resp.raise_for_status()

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "anomalies": anomalies,
                "invoiceStatus": invoice_status,
                "salesforceResponse": resp.json()
            }),
            status_code=200,
            mimetype="application/json"
        )

    except requests.exceptions.RequestException as e:
        logging.exception("Salesforce Invoice API call failed")
        return func.HttpResponse(
            json.dumps({
                "error": "Salesforce API error",
                "details": str(e)
            }),
            status_code=502,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Invoice analyze and sync failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

@app.route(route="projects/{projectName}/po-analyze-and-sync", methods=["POST"], auth_level=func.AuthLevel.ANONYMOUS)
def analyze_and_sync_po_to_salesforce(req: func.HttpRequest) -> func.HttpResponse:
    """
    Analyze PO against final report and sync to Salesforce in one operation.
    
    Request body:
    {
        "purchaseOrder": {
            "id": "PO-0001",
            "poNumber": "PO-0001",
            "vendor": "Tech Supplies Inc.",
            "receiptId": "RCP-0001",
            "createdBy": "John Doe",
            "createdAt": "2026-01-27",
            ...
        }
    }
    """
    logging.info("PO analyze and sync to Salesforce triggered")

    try:
        project = req.route_params.get("projectName")
        body = req.get_json()
        po = body.get("purchaseOrder")

        if not project or not po:
            return func.HttpResponse(
                json.dumps({"error": "projectName and purchaseOrder are required"}),
                status_code=400,
                mimetype="application/json"
            )

        # Step 1: Analyze PO against final report
        latest_final = get_latest_final_report_payload(project)
        final_content = latest_final["payload"].get("content", "")
        
        logging.info(f"Project: {project}")
        logging.info(f"PO Number: {po.get('poNumber')}")
        logging.info(f"Final Report Content Length: {len(final_content)}")
        logging.info(f"Final Report Content Preview: {final_content[:500]}")
        
        # Check if final content is empty
        if not final_content or final_content.strip() == "":
            logging.warning(f"WARNING: Final report content is empty for project {project}")
            return func.HttpResponse(
                json.dumps({
                    "error": f"Final report content not found for project {project}",
                    "poNumber": po.get("poNumber")
                }),
                status_code=400,
                mimetype="application/json"
            )

        # Log the PO data being analyzed
        logging.info(f"PO Data: {json.dumps(po, indent=2)}")

        existing_anomalies_raw = body.get("existingAnomalies") or po.get("existingAnomalies") or []

        prompt = f"""
You are a senior construction controls analyst.

TASK:
Compare the Purchase Order (PO) against the FINAL report content (SOW) and identify anomalies.

PO DATA (JSON):
{json.dumps(po)}

FINAL REPORT CONTENT:
{final_content}

OUTPUT FORMAT (STRICT JSON ARRAY):
[
  {{
    "id": "string",
    "type": "quantity_mismatch | price_mismatch | total_mismatch | scope_mismatch | timeline_mismatch | compliance_gap",
    "severity": "low | medium | high",
    "description": "ITEM NAME ONLY (no extra text)",
    "field": "string",
    "expectedValue": "string or number",
    "actualValue": "string or number",
    "step": "po_receipt"
  }}
]

RULES - CRITICAL:
- Return ONLY valid JSON (no markdown).
- If no anomalies, return EMPTY ARRAY: []
- Use step="po_receipt" so the frontend labels it as Step 1.
- Treat FINAL REPORT CONTENT as the SOW source of truth.
- ONLY flag actual mismatches where expectedValue ≠ actualValue.
- Do NOT flag items when expected == actual (they match perfectly).
- Only compare PO items to SOW items that match by description or equivalent name.
- For matched items, ONLY return an anomaly if quantity or price differs.
- If a PO item does NOT exist in SOW, return an anomaly (type: scope_mismatch).
- Do NOT flag SOW-only items that are not present in the PO.
- Do NOT compare or mention client/vendor names, dates, totals, project names, or any non-line-item fields.
- Be VERY STRICT: Only return an anomaly if there is a CLEAR NUMERIC MISMATCH.
- Do NOT return anomalies for items where values match exactly.
- Set description to ONLY the item name (e.g., Route Survey & Geotechnical Boreholes). Do not include any other text.

SEVERITY RULES:
- type: price_mismatch → severity: "high" (always)
- type: total_mismatch → severity: "high" (always)
- type: quantity_mismatch where PO_quantity > SOW_quantity → severity: "high" (ordering more than needed)
- type: quantity_mismatch where PO_quantity < SOW_quantity → severity: "medium" (ordering less than needed)
- type: scope_mismatch (PO item not in SOW) → severity: "high"
- if the quantity in PO is less than quantity in SOW -> severity: "medium"
"""

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You analyze PO vs final report and return anomalies as JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1
        )

        anomalies_raw = completion.choices[0].message.content.strip()
        logging.info(f"AI Response (Raw): {anomalies_raw}")
        
        try:
            anomalies = json.loads(anomalies_raw)
        except json.JSONDecodeError as e:
            logging.error(f"Failed to parse AI response as JSON: {str(e)}")
            logging.error(f"Raw response was: {anomalies_raw}")
            return func.HttpResponse(
                json.dumps({
                    "error": f"AI response was not valid JSON: {str(e)}",
                    "rawResponse": anomalies_raw
                }),
                status_code=500,
                mimetype="application/json"
            )
        
        # Use all anomalies from AI response (frontend handles deduplication with Created_At)
        logging.info(f"Parsed Anomalies: {json.dumps(anomalies, indent=2)}")
        logging.info(f"Total Anomalies Found: {len(anomalies)}")
        
        sf_auth = get_salesforce_access_token()
        access_token = sf_auth["access_token"]
        instance_url = sf_auth["instance_url"]

        # Format payload for Salesforce
        po_status = "To be Verified" if len(anomalies) > 0 else "Verified"
        
        # Convert anomalies to Salesforce format
        salesforce_anomalies = []
        
        for anomaly in anomalies:
            severity = (anomaly.get("severity") or "").lower()
            severity_label = {
                "high": "High Severity",
                "medium": "Medium Severity",
                "low": "Low Severity",
            }.get(severity, "Medium Severity")

            # Include field type in description for frontend parsing
            description = anomaly.get("description") or anomaly.get("type") or "Issue"
            field_type = anomaly.get("field", "amount")
            detail_with_type = f"{description} / {field_type}"

            salesforce_anomalies.append({
                "Detail": detail_with_type,
                "Severity": severity_label,
                "Expected": anomaly.get("expectedValue"),
                "Actual": anomaly.get("actualValue"),
                "Status": "Open"
            })

        # Resolve OrderID from incoming PO (sent from Salesforce API)
        order_id = po.get("OrderID") or po.get("orderId") or po.get("Order_Id") or po.get("Order_Id__c")

        if not order_id:
            return func.HttpResponse(
                json.dumps({
                    "error": "OrderID is required to update Salesforce",
                    "poNumber": po.get("poNumber")
                }),
                status_code=400,
                mimetype="application/json"
            )
        # If no anomalies found, fetch receipt details from Snowflake
        receipt_id = po.get("receiptId", "")
        created_by = po.get("createdBy", "")
        created_at = po.get("createdAt", "")
        receipt_data = None
        
        if len(anomalies) == 0:
            try:
                ctx = connection_snowflake("CONSTRUCTIONDEMO", "partner", "SCT_SARORA_DB")
                cursor = ctx.cursor()
                
                # Fetch receipt details by OrderID including line items
                cursor.execute("SELECT RECEIPT_ID, CREATED_BY, CREATED_AT, LIST_ITEMS, STATUS FROM RECEIPTS WHERE ORDER_ID = %s LIMIT 1", [order_id])
                result = cursor.fetchone()
                
                if result:
                    receipt_id = result[0] or ""
                    created_by = result[1] or ""
                    created_at = str(result[2]) if result[2] else ""
                    list_items_raw = result[3]
                    receipt_status = result[4] or ""
                    
                    # Parse line items if they exist
                    line_items = []
                    if list_items_raw:
                        try:
                            if isinstance(list_items_raw, str):
                                line_items = json.loads(list_items_raw)
                            elif isinstance(list_items_raw, list):
                                line_items = list_items_raw
                            else:
                                # Snowflake might return it as a dict or other type
                                line_items = list_items_raw if list_items_raw else []
                        except Exception as e:
                            logging.warning(f"Failed to parse line items: {str(e)}")
                            line_items = []
                    
                    receipt_data = {
                        "receiptId": receipt_id,
                        "createdBy": created_by,
                        "createdAt": created_at,
                        "lineItems": line_items,
                        "receiptstatus": receipt_status
                    }
                    logging.info(f"Fetched receipt data from Snowflake: {receipt_data}")
                
                disconnect_snowflake(ctx)
            except Exception as e:
                logging.warning(f"Failed to fetch receipt from Snowflake: {str(e)}")

        payload = {
            "OrderID": order_id,
            "Client_Name": po.get("vendor", "Unknown"),
            "Receipt_ID": receipt_id,
            "Created_By": created_by,
            "Created_At": created_at,
            "POStatus": po_status,
            "Anomalies": salesforce_anomalies
        }

        # Wrap payload in array for Salesforce API
        payload_array = [payload]
        
        logging.info(f"Payload: {json.dumps(payload_array, indent=2)}")

        # Call Salesforce API
        salesforce_url = instance_url.rstrip("/") + "/services/apexrest/PurchaseOrder"
        logging.info(f"Calling Salesforce API: {salesforce_url}")

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        resp = requests.post(salesforce_url, json=payload_array, headers=headers)
        
        # Log response for debugging
        logging.info(f"Salesforce API Response Status: {resp.status_code}")
        logging.info(f"Salesforce API Response: {resp.text}")
        
        resp.raise_for_status()

        # Parse Salesforce response to check if PO is verified
        try:
            salesforce_response = resp.json()
            logging.info(f"Parsed Salesforce response: {salesforce_response}")
            
            # Check if response indicates Verified status and we haven't fetched receipt yet
            if isinstance(salesforce_response, list) and len(salesforce_response) > 0:
                sf_status = salesforce_response[0].get("POStatus") or salesforce_response[0].get("status")
                logging.info(f"Salesforce PO Status: {sf_status}")
                
                if sf_status == "Verified" and not receipt_data:
                    logging.info(f"PO is Verified in Salesforce, fetching receipt data from Snowflake for OrderID: {order_id}")
                    try:
                        ctx = connection_snowflake("CONSTRUCTIONDEMO", "partner", "SCT_SARORA_DB")
                        cursor = ctx.cursor()
                        
                        cursor.execute("SELECT RECEIPT_ID, CREATED_BY, CREATED_AT, LIST_ITEMS, STATUS FROM RECEIPTS WHERE ORDER_ID = %s LIMIT 1", [order_id])
                        result = cursor.fetchone()
                        
                        if result:
                            receipt_id = result[0] or ""
                            created_by = result[1] or ""
                            created_at = str(result[2]) if result[2] else ""
                            list_items_raw = result[3]
                            receipt_status = result[4] or ""
                            
                            line_items = []
                            if list_items_raw:
                                try:
                                    if isinstance(list_items_raw, str):
                                        line_items = json.loads(list_items_raw)
                                    elif isinstance(list_items_raw, list):
                                        line_items = list_items_raw
                                    else:
                                        line_items = list_items_raw if list_items_raw else []
                                except Exception as e:
                                    logging.warning(f"Failed to parse line items: {str(e)}")
                                    line_items = []
                            
                            receipt_data = {
                                "receiptId": receipt_id,
                                "createdBy": created_by,
                                "createdAt": created_at,
                                "lineItems": line_items,
                                "receiptstatus": receipt_status
                            }
                            logging.info(f"Fetched receipt data after Salesforce verification: {receipt_data}")
                        
                        disconnect_snowflake(ctx)
                    except Exception as e:
                        logging.warning(f"Failed to fetch receipt after Salesforce verification: {str(e)}")
        except Exception as e:
            logging.warning(f"Failed to parse Salesforce response: {str(e)}")

        logging.info(f"Successfully analyzed and synced PO to Salesforce: {po.get('poNumber')}")

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": f"PO analyzed and synced to Salesforce with status: {po_status}",
                "poNumber": po.get("poNumber"),
                "anomalies": anomalies,
                "anomalyCount": len(anomalies),
                "status": po_status,
                "salesforcePayload": payload,
                "receiptData": receipt_data
            }),
            status_code=200,
            mimetype="application/json"
        )

    except requests.HTTPError as e:
        logging.exception("Salesforce API error")
        error_details = {
            "error": "Salesforce API failed",
            "status_code": e.response.status_code if hasattr(e, 'response') else None,
            "details": str(e),
            "response_text": e.response.text if hasattr(e, 'response') else None
        }
        return func.HttpResponse(
            json.dumps(error_details),
            status_code=502,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("PO analyze and sync failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )

@app.route(route="projects/{projectName}/finalize", methods=["POST"])
def finalize_document(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Finalize document triggered")

    project = req.route_params.get("projectName")
    body = req.get_json()

    # ✅ align with frontend
    file_name = body.get("finalFile")

    logging.info(f"Finalizing document | Project={project} | File={file_name}")

    if not project or not file_name:
        return func.HttpResponse(
            json.dumps({"error": "projectName and finalFile required"}),
            status_code=400,
            mimetype="application/json"
        )

    # 1️⃣ Read final PDF
    extracted_text = read_pdf_from_blob(project, file_name)

    payload = {
        "project": project,
        "fileName": file_name,
        "finalizedAt": datetime.utcnow().isoformat(),
        "content": extracted_text
    }

    excel_path = f"daily-reports/{project}/final-report.xlsx"
    blob_client = get_blob_client(excel_path)
    output = io.BytesIO()

    try:
        existing = io.BytesIO()
        blob_client.download_blob().readinto(existing)
        existing.seek(0)
        wb = load_workbook(existing)
        ws = wb.active
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.title = "final_logs"
        ws.append(["date", "type", "data"])

    ws.append([
        datetime.utcnow().strftime("%Y-%m-%d"),
        "final",
        json.dumps(payload)
    ])

    wb.save(output)
    output.seek(0)

    blob_client.upload_blob(
        output,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )

    return func.HttpResponse(
        json.dumps({
            "success": True,
            "excelPath": excel_path
        }),
        status_code=200,
        mimetype="application/json"
    )


def normalize_date(value):
    """
    Handles Excel datetime, date, and string formats safely
    """
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        # Handles both "YYYY-MM-DD" and "YYYY-MM-DD HH:MM:SS"
        return datetime.fromisoformat(value).date()

    raise ValueError(f"Unsupported date format: {value}")


@app.route(route="projects/{projectName}/progress-chart", methods=["GET"])
def generate_progress_chart(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Generating project progress chart")

    try:
        project_name = req.route_params.get("projectName")
        start_date_str = req.params.get("startDate")

        if not project_name:
            return func.HttpResponse(
                json.dumps({"error": "projectName is required"}),
                status_code=400,
                mimetype="application/json"
            )

        if not start_date_str:
            return func.HttpResponse(
                json.dumps({"error": "startDate is required (YYYY-MM-DD)"}),
                status_code=400,
                mimetype="application/json"
            )

        project_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()

        # ---------- Download final report ----------
        excel_path = f"daily-reports/{project_name}/final-report.xlsx"
        blob_client = get_blob_client(excel_path)

        stream = io.BytesIO()
        blob_client.download_blob().readinto(stream)
        stream.seek(0)

        wb = load_workbook(stream)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        headers = [h.lower() for h in rows[0]]
        data_rows = rows[1:]

        daily_reports = []
        final_report_data = None

        for row in data_rows:
            row_dict = dict(zip(headers, row))

            if row_dict.get("type") == "daily_report":
                row_dict["normalized_date"] = normalize_date(row_dict["date"])
                daily_reports.append(row_dict)

            elif row_dict.get("type") == "final":
                final_report_data = row_dict.get("data")

        if not daily_reports:
            return func.HttpResponse(
                json.dumps({"error": "No daily reports found"}),
                status_code=400,
                mimetype="application/json"
            )

        if not final_report_data:
            return func.HttpResponse(
                json.dumps({"error": "No final report found"}),
                status_code=400,
                mimetype="application/json"
            )

        # ---------- Sort daily reports ----------
        daily_reports.sort(key=lambda x: x["normalized_date"])

        # ---------- Prepare LLM Prompt ----------
        daily_texts = "\n\n".join([
            f"Date: {d['normalized_date'].strftime('%Y-%m-%d')}\nDaily Report:\n{d['data']}"
            for d in daily_reports
        ])

        prompt = f"""
You are a senior construction project analyst.

FINAL REPORT:
{final_report_data}

DAILY REPORTS:
{daily_texts}

TASK:
For each daily report, compare the activities with the FINAL report and estimate cumulative progress.

Return ONLY valid JSON in this format:
[
  {{"date": "YYYY-MM-DD", "progress": number}}
]
"""

        completion = client.chat.completions.create(
            model=DEPLOYMENT_NAME,
            messages=[
                {"role": "system", "content": "You analyze construction project progress."},
                {"role": "user", "content": prompt}
            ],
            temperature=0
        )

        llm_response = completion.choices[0].message.content.strip()
        llm_data = json.loads(llm_response)

        # ---------- Build progress map ----------
        progress_map = {
            datetime.fromisoformat(item["date"]).date(): item["progress"]
            for item in llm_data
        }

        # ---------- Fill missing dates from project start ----------
        end_date = daily_reports[-1]["normalized_date"]
        current_date = project_start_date

        chart_data = []
        last_progress = 0

        while current_date <= end_date:
            if current_date in progress_map:
                last_progress = progress_map[current_date]

            chart_data.append({
                "date": current_date.strftime("%Y-%m-%d"),
                "progress": last_progress
            })

            current_date += timedelta(days=1)

        return func.HttpResponse(
            json.dumps({"project": project_name,"chartData": chart_data}),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Failed to generate progress chart")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )
    

def get_salesforce_access_token():
    """Fetch Salesforce access token using refresh_token flow"""
    token_url = os.getenv("SF_TOKEN_URL")

    payload = {
        "grant_type": "refresh_token",
        "client_id": os.getenv("SF_ClientID"),
        "client_secret": os.getenv("SF_clientSecret"),
        "refresh_token": os.getenv("SF_refresh_token"),
    }

    headers = {
        "Content-Type": "application/x-www-form-urlencoded"
    }

    resp = requests.post(token_url, data=payload, headers=headers)
    resp.raise_for_status()

    return resp.json()


def fetch_receipt_data_by_order_id(order_id: str, project_name: str = None):
    """Fetch receipt data from Snowflake by OrderID including anomalies. 
    If receiptstatus is 'submitted', also fetch invoice data from Salesforce."""
    if not order_id:
        return None

    ctx = None
    try:
        ctx = connection_snowflake("CONSTRUCTIONDEMO", "partner", "SCT_SARORA_DB")
        cursor = ctx.cursor()

        cursor.execute(
            "SELECT RECEIPT_ID, CREATED_BY, CREATED_AT, LIST_ITEMS, ANOMALIES, STATUS FROM RECEIPTS WHERE ORDER_ID = %s LIMIT 1",
            [order_id]
        )
        result = cursor.fetchone()

        if not result:
            return None

        receipt_id = result[0] or ""
        created_by = result[1] or ""
        created_at = str(result[2]) if result[2] else ""
        list_items_raw = result[3]
        anomalies_raw = result[4]
        receipt_status = result[5] or ""

        line_items = []
        if list_items_raw:
            try:
                if isinstance(list_items_raw, str):
                    line_items = json.loads(list_items_raw)
                elif isinstance(list_items_raw, list):
                    line_items = list_items_raw
                else:
                    line_items = list_items_raw if list_items_raw else []
            except Exception as e:
                logging.warning(f"Failed to parse line items: {str(e)}")
                line_items = []

        anomalies = []
        if anomalies_raw:
            try:
                if isinstance(anomalies_raw, str):
                    anomalies = json.loads(anomalies_raw)
                elif isinstance(anomalies_raw, list):
                    anomalies = anomalies_raw
                else:
                    anomalies = anomalies_raw if anomalies_raw else []
            except Exception as e:
                logging.warning(f"Failed to parse anomalies: {str(e)}")
                anomalies = []

        receipt_data = {
            "receiptId": receipt_id,
            "createdBy": created_by,
            "createdAt": created_at,
            "lineItems": line_items,
            "anomalies": anomalies,
            "receiptstatus": receipt_status
        }

        # If receiptstatus is "submitted", fetch invoice data from Salesforce
        if receipt_status and receipt_status.lower() == "submitted":
            try:
                logging.info(f"Receipt status is 'submitted', fetching invoice data for OrderID: {order_id}")
                sf_auth = get_salesforce_access_token()
                access_token = sf_auth["access_token"]
                instance_url = sf_auth["instance_url"]

                # Call Salesforce Invoice API
                invoice_url = instance_url.rstrip("/") + "/services/apexrest/Invoice"
                
                params = {
                    "projectName": project_name or "default",
                    "orderId": order_id
                }

                headers = {
                    "Authorization": f"Bearer {access_token}",
                    "Content-Type": "application/json"
                }

                resp = requests.get(invoice_url, headers=headers, params=params)
                resp.raise_for_status()

                invoice_data = resp.json()
                logging.info(f"Successfully fetched invoice data for OrderID {order_id}")
                receipt_data["invoiceData"] = invoice_data
            except Exception as e:
                logging.warning(f"Failed to fetch invoice data for OrderID {order_id}: {str(e)}")

        return receipt_data
    except Exception as e:
        logging.warning(f"Failed to fetch receipt by OrderID {order_id}: {str(e)}")
        return None
    finally:
        if ctx:
            try:
                disconnect_snowflake(ctx)
            except Exception:
                pass


@app.route(
    route="projects/{projectName}/invoices",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def get_project_invoices(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Fetching invoices via Salesforce")

    try:
        project_name = req.route_params.get("projectName")
        if not project_name:
            return func.HttpResponse(
                json.dumps({"error": "projectName required"}),
                status_code=400,
                mimetype="application/json"
            )

        # 1️⃣ Get Salesforce token
        sf_auth = get_salesforce_access_token()
        logging.info("Obtained Salesforce access token")
        access_token = sf_auth["access_token"]
        logging.info(access_token)
        instance_url = sf_auth["instance_url"]
        logging.info(instance_url)

        # 2️⃣ Call Salesforce invoices API
        # invoice_api_path = os.getenv("SF_INVOICE_API")
        invoice_url = instance_url.rstrip("/") + "/services/apexrest/PurchaseOrder"
        logging.info(f"Calling Salesforce Invoice API: {invoice_url}")

        params = {
            "projectName": project_name
        }

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        resp = requests.get(invoice_url, headers=headers, params=params)
        resp.raise_for_status()

        invoices = resp.json()

        # If Salesforce returns a list, attach receipt data when PO_Status is Verified
        if isinstance(invoices, list):
            for invoice in invoices:
                po_status = (
                    invoice.get("PO_Status")
                    or invoice.get("POStatus")
                    or invoice.get("status")
                )
                if (po_status or "").lower() == "verified":
                    order_id = (
                        invoice.get("OrderID")
                        or invoice.get("orderId")
                        or invoice.get("Order_Id")
                        or invoice.get("Order_Id__c")
                    )
                    if order_id:
                        receipt_data = fetch_receipt_data_by_order_id(order_id, project_name)
                        if receipt_data:
                            invoice["receiptData"] = receipt_data

        return func.HttpResponse(
            json.dumps(invoices),
            status_code=200,
            mimetype="application/json"
        )

    except requests.HTTPError as e:
        logging.exception("Salesforce API error")
        return func.HttpResponse(
            json.dumps({
                "error": "Salesforce API failed",
                "details": str(e)
            }),
            status_code=502,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Invoice fetch failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(route="projects/{projectName}/invoice/{orderId}", methods=["GET"], auth_level=func.AuthLevel.ANONYMOUS)
def get_invoice_by_order_id(req: func.HttpRequest) -> func.HttpResponse:
    """
    Fetch a single invoice from Salesforce by OrderID.
    """
    try:
        project_name = req.route_params.get("projectName")
        order_id = req.route_params.get("orderId")

        if not project_name or not order_id:
            return func.HttpResponse(
                json.dumps({"error": "projectName and orderId are required"}),
                status_code=400,
                mimetype="application/json"
            )

        logging.info(f"Fetching invoice for project: {project_name}, orderId: {order_id}")

        # Get Salesforce token
        sf_auth = get_salesforce_access_token()
        access_token = sf_auth["access_token"]
        instance_url = sf_auth["instance_url"]

        # Call Salesforce Invoice API
        invoice_url = instance_url.rstrip("/") + "/services/apexrest/Invoice"
        logging.info(f"Calling Salesforce Invoice API: {invoice_url}")

        params = {
            "projectName": project_name,
            "orderId": order_id
        }

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        resp = requests.get(invoice_url, headers=headers, params=params)
        resp.raise_for_status()

        invoice_data = resp.json()
        logging.info(f"Retrieved invoice for orderId {order_id}: {json.dumps(invoice_data, indent=2)}")

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "invoice": invoice_data
            }),
            status_code=200,
            mimetype="application/json"
        )

    except requests.exceptions.RequestException as e:
        logging.exception("Salesforce Invoice API call failed")
        return func.HttpResponse(
            json.dumps({
                "error": "Salesforce API error",
                "details": str(e)
            }),
            status_code=502,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Invoice fetch by orderId failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )



# ============== SNOWFLAKE CONNECTION ===============

def connection_snowflake(schema, environment, database):
    """
    Connect to Snowflake using Key Pair authentication.
    Note: Consider adding Key Vault integration to retrieve secrets from cloud instead of hardcoding.
    
    Args:
        schema: Snowflake schema name
        environment: Environment type ('training' or 'partner')
        database: Database name
    
    Returns:
        Snowflake connector object
    """
    try:
        # Load private key from file (TODO: Replace with Key Vault retrieval)
        current_directory = os.path.dirname(os.path.abspath(__file__))
        key_path = os.path.join(current_directory, '', 'rsa_key.p8')
        
        with open(key_path, "rb") as key:
            p_key = serialization.load_pem_private_key(
                key.read(),
                password='123'.encode(),
                backend=default_backend()
            )
        
        pkb = p_key.private_bytes(
            encoding=serialization.Encoding.DER,
            format=serialization.PrivateFormat.PKCS8,
            encryption_algorithm=serialization.NoEncryption()
        )
        
        if environment == 'training':
            ctx = snowflake.connector.connect(
                user='sanchit.arora@cginfinity.com',
                account='ap34643.east-us-2.azure',
                private_key=pkb,
                warehouse='SCT_SARORA_WH',
                database='sct_sarora_db',
                schema=schema
            )
        elif environment == 'partner':
            ctx = snowflake.connector.connect(
                user='sanchit.arora@cginfinity.com',
                account='VLB03298-CGINFINITY_PARTNER',
                private_key=pkb,
                warehouse='SCT_SARORA_WH',
                database='sct_sarora_db',
                schema=schema
            )
        else:
            raise ValueError("Invalid environment specified. Choose 'training' or 'partner'.")
        
        logging.info(f"Successfully connected to Snowflake - Environment: {environment}, Database: {database}, Schema: {schema}")
        return ctx
    
    except Exception as e:
        logging.exception("Failed to connect to Snowflake")
        raise


def disconnect_snowflake(conn):
    """Close Snowflake connection"""
    try:
        conn.close()
        logging.info("Snowflake connection closed")
    except Exception as e:
        logging.exception("Failed to disconnect from Snowflake")
        raise


@app.route(
    route="snowflake/query",
    methods=["POST"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def snowflake_query(req: func.HttpRequest) -> func.HttpResponse:
    """
    Execute a query on Snowflake.
    
    Request body:
    {
        "query": "SELECT * FROM table",
        "environment": "training" or "partner",
        "database": "",
        "schema": "schema_name"
    }
    """
    try:
        req_body = req.get_json()
        query = "Select * from Insights.Demographics_stats LIMIT 10"  #req_body.get("query")
        environment = req_body.get("environment", "training")
        database = req_body.get("database", "SCT_SARORA_DB")
        schema = "Insights"
        
        if not query or not schema:
            return func.HttpResponse(
                json.dumps({"error": "query and schema are required"}),
                status_code=400,
                mimetype="application/json"
            )
        
        # Connect to Snowflake
        ctx = connection_snowflake(schema, environment, database)
        cursor = ctx.cursor()
        
        # Execute query
        cursor.execute(query)
        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        
        # Convert results to list of dictionaries
        data = []
        for row in results:
            data.append(dict(zip(columns, row)))
        
        disconnect_snowflake(ctx)
        
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "data": data,
                "rowCount": len(data)
            }),
            status_code=200,
            mimetype="application/json"
        )
    
    except Exception as e:
        logging.exception("Snowflake query failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(
    route="snowflake/test-connection",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def snowflake_test_connection(req: func.HttpRequest) -> func.HttpResponse:
    """
    Test Snowflake connection with specified environment.
    
    Query parameters:
    - environment: 'training' or 'partner' (default: 'training')
    """
    try:
        environment = req.params.get("environment", "")
        
        # Test connection
        ctx = connection_snowflake("PUBLIC", environment, "Marketplace")
        
        # Simple query to test connection
        cursor = ctx.cursor()
        cursor.execute("SELECT CURRENT_USER(), CURRENT_WAREHOUSE(), CURRENT_DATABASE()")
        result = cursor.fetchone()
        
        disconnect_snowflake(ctx)
        
        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": "Successfully connected to Snowflake",
                "user": result[0],
                "warehouse": result[1],
                "database": result[2]
            }),
            status_code=200,
            mimetype="application/json"
        )
    
    except Exception as e:
        logging.exception("Snowflake connection test failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(
    route="receipts",
    methods=["GET"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def get_receipts(req: func.HttpRequest) -> func.HttpResponse:
    """
    Fetch receipts from Snowflake RECEIPTS table.

    Query parameters:
    - orderId: optional order id filter
    - receiptId: optional receipt id filter
    - environment: 'training' or 'partner' (default: 'training')
    - database: database name (default: 'Marketplace')
    - schema: schema name (default: 'PUBLIC')
    - limit: max rows (default: 100)
    """
    try:
        order_id = req.params.get("orderId")
        receipt_id = req.params.get("receiptId")
        environment = req.params.get("environment", "partner")
        database = req.params.get("database", "SCT_SARORA_DB")
        schema = req.params.get("schema", "CONSTRUCTIONDEMO")
        # limit = req.params.get("limit", "100")

        sql = f"SELECT * FROM RECEIPTS"
        conditions = []
        params = []

        if order_id:
            conditions.append("ORDER_ID = %s")
            params.append(order_id.upper())

        if receipt_id:
            conditions.append("RECEIPT_ID = %s")
            params.append(receipt_id)

        if conditions:
            sql += " WHERE " + " AND ".join(conditions)

        sql += " ORDER BY UPDATED_AT DESC"
        logging.info(f"Executing SQL: {sql} with params: {params}")
        ctx = connection_snowflake(schema, environment, database)
        cursor = ctx.cursor()

        if params:
            cursor.execute(sql, params)
        else:
            cursor.execute(sql)

        results = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        def serialize_value(value):
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            return value

        data = []
        for row in results:
            row_dict = {columns[i]: serialize_value(row[i]) for i in range(len(columns))}

            list_items = row_dict.get("LIST_ITEMS")
            if isinstance(list_items, str):
                try:
                    list_items = json.loads(list_items)
                except Exception:
                    pass

            anomalies = row_dict.get("ANOMALIES")
            if isinstance(anomalies, str):
                try:
                    anomalies = json.loads(anomalies)
                except Exception:
                    anomalies = []
            elif not anomalies:
                anomalies = []

            data.append({
                "receiptId": row_dict.get("RECEIPT_ID"),
                "orderId": row_dict.get("ORDER_ID"),
                "createdAt": row_dict.get("CREATED_AT"),
                "updatedAt": row_dict.get("UPDATED_AT"),
                "createdBy": row_dict.get("CREATED_BY"),
                "status": row_dict.get("STATUS"),
                "lineItems": list_items or [],
                "anomalies": anomalies or []
            })
            logging.info(f"Processed receipt: {row_dict.get('RECEIPT_ID')}")

        disconnect_snowflake(ctx)

        return func.HttpResponse(
            json.dumps({"success": True, "data": data, "rowCount": len(data)}),
            status_code=200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Fetch receipts failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )


@app.route(
    route="receipts",
    methods=["POST"],
    auth_level=func.AuthLevel.ANONYMOUS
)
def create_or_update_receipt(req: func.HttpRequest) -> func.HttpResponse:
    """
    Create or update a receipt in Snowflake RECEIPTS table.

    Request body:
    {
        "receiptId": "RCPT-001",
        "orderId": "ORD-1001",
        "status": "draft" or "approved" or "rejected",
        "createdBy": "site_manager_1",
        "lineItems": [
            {
                "item_id": "ITM-01",
                "item_name": "Cement Bags",
                "quantity": 50,
                "unit_price": 380,
                "total_price": 19000,
                "unit": "bags"
            }
        ],
        "environment": "training" or "partner" (optional),
        "database": "database_name" (optional),
        "schema": "schema_name" (optional)
    }
    """
    try:
        req_body = req.get_json()
        
        # Extract fields
        receipt_id = req_body.get("receiptId")
        order_id = req_body.get("orderId")
        status = req_body.get("status", "draft")
        created_by = req_body.get("createdBy")
        line_items = req_body.get("lineItems", [])
        environment = req_body.get("environment", "partner")
        database = req_body.get("database", "SCT_SARORA_DB")
        schema = req_body.get("schema", "CONSTRUCTIONDEMO")

        # Validation
        if not receipt_id or not order_id or not created_by:
            return func.HttpResponse(
                json.dumps({"error": "receiptId, orderId, and createdBy are required"}),
                status_code=400,
                mimetype="application/json"
            )

        # Validate status
        valid_statuses = ['draft', 'approved', 'rejected']
        if status not in valid_statuses:
            return func.HttpResponse(
                json.dumps({"error": f"Invalid status. Must be one of: {', '.join(valid_statuses)}"}),
                status_code=400,
                mimetype="application/json"
            )

        # Convert line items to JSON string
        line_items_json = json.dumps(line_items) if line_items else "[]"

        # Connect to Snowflake
        ctx = connection_snowflake(schema, environment, database)
        cursor = ctx.cursor()

        # Check if receipt already exists
        cursor.execute(
            "SELECT RECEIPT_ID FROM RECEIPTS WHERE RECEIPT_ID = %s",
            [receipt_id]
        )
        existing = cursor.fetchone()

        if existing:
            # Update existing receipt
            update_sql = (
                "UPDATE RECEIPTS "
                "SET ORDER_ID = %s, STATUS = %s, CREATED_BY = %s, LIST_ITEMS = TRY_PARSE_JSON(%s), UPDATED_AT = CURRENT_TIMESTAMP() "
                "WHERE RECEIPT_ID = %s"
            )
            cursor.execute(update_sql, [order_id, status, created_by, line_items_json, receipt_id])
            action = "updated"
        else:
            # Insert new receipt using SELECT to handle PARSE_JSON
            insert_sql = (
                "INSERT INTO RECEIPTS (RECEIPT_ID, ORDER_ID, STATUS, CREATED_BY, LIST_ITEMS) "
                "SELECT %s, %s, %s, %s, TRY_PARSE_JSON(%s)"
            )
            cursor.execute(insert_sql, [receipt_id, order_id, status, created_by, line_items_json])
            action = "created"

        # Commit changes
        ctx.commit()
        disconnect_snowflake(ctx)

        logging.info(f"Receipt {action} successfully: {receipt_id}")

        return func.HttpResponse(
            json.dumps({
                "success": True,
                "message": f"Receipt {action} successfully",
                "receiptId": receipt_id,
                "orderId": order_id,
                "action": action
            }),
            status_code=201 if action == "created" else 200,
            mimetype="application/json"
        )

    except Exception as e:
        logging.exception("Create/update receipt failed")
        return func.HttpResponse(
            json.dumps({"error": str(e)}),
            status_code=500,
            mimetype="application/json"
        )